import tkinter as tk
from tkinter import ttk
import datetime
from tkinter.constants import ACTIVE
import openpyxl as xl
import pandas as pd


class TrieNode:
    def __init__(self):
        self.children = {}
        self.is_end = False
        self.word = None


class Trie:
    def __init__(self):
        self.root = TrieNode()

    def insert(self, word):
        node = self.root
        for char in word.lower():
            if char not in node.children:
                node.children[char] = TrieNode()
            node = node.children[char]
        node.is_end = True
        node.word = word  # store original word

    def search_prefix(self, prefix, limit=20):
        node = self.root
        for char in prefix.lower():
            if char not in node.children:
                return []  # no matches
            node = node.children[char]

        results = []
        self._dfs(node, results, limit)
        return results

    def _dfs(self, node, results, limit):
        if len(results) >= limit:
            return
        if node.is_end:
            results.append(node.word)
        for child in node.children.values():
            self._dfs(child, results, limit)


class PrescriptionApp:
    def __init__(self, root):
        self.window = root  # Fixed: was 'window' instead of 'root'
        self.drug_entries = []
        self.drug_counter = 0
        self.current_search_widget = None  # Track which widget is currently being searched
        self.window.title("Make Prescription")
        self.create_main_frame()
        self.create_user_info_frame()
        self.drug_list_view_frame()  # Move this before create_drug_info_frame
        self.create_drug_info_frame()
        self.add_drug_entry()
        self.create_button_frame()

    def create_main_frame(self):
        self.frame = tk.Frame(self.window)
        self.frame.pack()

    def create_excel_view_frame(self):
        self.excel_view_frame = ttk.Frame(self.frame)
        self.excel_view_frame.grid(row=0, column=1, sticky='nsew')
        treeScroll = ttk.Scrollbar(self.excel_view_frame, orient='vertical')
        treeScroll.pack(side='right', fill='y')
        cols = ("Name", "Age", "Sex")
        self.treeView = ttk.Treeview(self.excel_view_frame, columns=cols, show="headings",
                                     yscrollcommand=treeScroll.set)

        self.treeView.pack(fill='both', expand=True)
        treeScroll.config(command=self.treeView.yview)

        path = "Sample.xlsx"
        wb = xl.load_workbook(path)
        sheet = wb.active

        list_values = list(sheet.values)
        print(list_values)
        for col in list_values[0]:
            self.treeView.heading(col, text=col)

        for values in list_values[1:]:
            self.treeView.insert(parent='', index='end', values=values)

    def insert_patient_info(self):
        name = self.patient_name_entry.get()
        age = self.patient_age_entry.get()
        sex = self.patient_sex_entry.get()

        path = "Sample.xlsx"
        wb = xl.load_workbook(path)
        sheet = wb.active

        row_values = [name, age, sex]
        sheet.append(row_values)
        wb.save(path)
        self.treeView.insert(parent='', index='end', values=row_values)

    def create_user_info_frame(self):
        self.user_info_frame = tk.LabelFrame(self.frame, text="User Information", width=300, height=200)
        self.user_info_frame.grid(row=0, column=0, padx=10, pady=10)

        # Date
        cur_date = tk.Label(self.user_info_frame, text="Date")
        cur_date.grid(row=0, column=0)
        self.cur_date_entry = tk.Entry(self.user_info_frame)
        self.cur_date_entry.insert(0, datetime.date.today())
        self.cur_date_entry.grid(row=0, column=1)

        # Doctor Name
        doctor_name = tk.Label(self.user_info_frame, text="Doctor Name:")
        doctor_name.grid(row=0, column=2)
        self.doctor_name_entry = tk.Entry(self.user_info_frame)
        self.doctor_name_entry.insert(0, "Madhavi Sapre")
        self.doctor_name_entry.grid(row=0, column=3)

        # Patient Name
        self.patient_name = tk.Label(self.user_info_frame, text="Patient Name:")
        self.patient_name.grid(row=1, column=0)
        self.patient_name_entry = tk.Entry(self.user_info_frame)
        self.patient_name_entry.grid(row=1, column=1)

        # Patient Age
        patient_age = tk.Label(self.user_info_frame, text="Patient Age:")
        patient_age.grid(row=1, column=2)
        self.patient_age_entry = ttk.Spinbox(self.user_info_frame, from_=0, to=100)
        self.patient_age_entry.grid(row=1, column=3)

        # Patient Sex
        patient_sex = tk.Label(self.user_info_frame, text="Patient Sex:")
        patient_sex.grid(row=2, column=0)
        self.patient_sex_entry = ttk.Combobox(self.user_info_frame, values=["Male", "Female", "Other"])
        self.patient_sex_entry.grid(row=2, column=1)

        # Patient Diagnosis
        patient_diagnosis = tk.Label(self.user_info_frame, text="Diagnosis:")
        patient_diagnosis.grid(row=3, column=0)
        self.patient_diagnosis_entry = tk.Entry(self.user_info_frame)
        self.patient_diagnosis_entry.grid(row=3, column=1)

    def create_button_frame(self):
        self.buttons_frame = tk.LabelFrame(self.frame)
        self.buttons_frame.grid(row=2, column=0, padx=10, pady=10)

        add_to_excel = tk.Button(self.buttons_frame, text="Add to Excel", command=self.save_prescription_to_excel)
        add_to_excel.grid(row=0, column=0, padx=10, pady=10)

    def create_drug_info_frame(self):
        self.drug_info_frame = tk.LabelFrame(self.frame, text="Drug Information")
        self.drug_info_frame.grid(row=1, column=0, padx=10, pady=10)

        tk.Label(self.drug_info_frame, text="Drug Name", font=("Arial", 10, "bold")).grid(row=0, column=0, padx=5,
                                                                                          pady=5)
        tk.Label(self.drug_info_frame, text="Dosage", font=("Arial", 10, "bold")).grid(row=0, column=1, padx=5, pady=5)
        tk.Label(self.drug_info_frame, text="Frequency", font=("Arial", 10, "bold")).grid(row=0, column=4, padx=5,
                                                                                          pady=5)
        tk.Label(self.drug_info_frame, text="Duration", font=("Arial", 10, "bold")).grid(row=0, column=6, padx=5,
                                                                                         pady=5)
        tk.Label(self.drug_info_frame, text="Instructions", font=("Arial", 10, "bold")).grid(row=0, column=8, padx=5,
                                                                                             pady=5)
        tk.Label(self.drug_info_frame, text="Action", font=("Arial", 10, "bold")).grid(row=0, column=9, padx=5, pady=5)

        self.drug_entries_frame = tk.Frame(self.drug_info_frame)
        self.drug_entries_frame.grid(row=1, column=0, columnspan=9)

        self.add_drug_btn = tk.Button(self.drug_info_frame, text="+ Add Drug", command=self.add_drug_entry)
        self.add_drug_btn.grid(row=2, column=4, pady=10)

    def add_drug_entry(self):
        row = len(self.drug_entries)

        drug_name = tk.Entry(self.drug_entries_frame, width=20)
        dosage = tk.Entry(self.drug_entries_frame, width=15)
        breakfast = ttk.Checkbutton(self.drug_entries_frame, text="Breakfast")
        lunch = ttk.Checkbutton(self.drug_entries_frame, text="Lunch")
        dinner = ttk.Checkbutton(self.drug_entries_frame, text="Dinner")
        after_meals = ttk.Radiobutton(self.drug_entries_frame, text="After Meals")
        before_meals = ttk.Radiobutton(self.drug_entries_frame, text="Before Meals")
        duration = tk.Entry(self.drug_entries_frame, width=15)
        instructions = tk.Entry(self.drug_entries_frame, width=25)
        remove_btn = tk.Button(self.drug_entries_frame, text="Remove", command=lambda r=row: self.remove_drug_entry(r))

        drug_name.grid(row=row, column=0, padx=5, pady=2)
        dosage.grid(row=row, column=1, padx=5, pady=2)
        breakfast.grid(row=row, column=2, padx=5, pady=2)
        lunch.grid(row=row, column=3, padx=5, pady=2)
        dinner.grid(row=row, column=4, padx=5, pady=2)
        after_meals.grid(row=row, column=5, padx=5, pady=2)
        before_meals.grid(row=row, column=6, padx=5, pady=2)
        duration.grid(row=row, column=7, padx=5, pady=2)
        instructions.grid(row=row, column=8, padx=5, pady=2)
        remove_btn.grid(row=row, column=9, padx=5, pady=2)

        drug_entry = {
            'drug_name': drug_name,
            'dosage': dosage,
            'breakfast': breakfast,
            'lunch': lunch,
            'dinner': dinner,
            'after_meals': after_meals,
            'before_meals': before_meals,
            'duration': duration,
            'instructions': instructions,
            'remove_btn': remove_btn,
            'row': row
        }

        # Bind events to the new drug name entry
        drug_name.bind('<KeyRelease>', self.check)
        drug_name.bind('<FocusIn>', lambda e, widget=drug_name: self.set_current_widget(widget))

        self.drug_entries.append(drug_entry)
        self.drug_counter += 1
        self.window.update_idletasks()

    def set_current_widget(self, widget):
        """Set the current widget that should receive autocomplete updates"""
        self.current_search_widget = widget

    def remove_drug_entry(self, row_index):
        # Validate row_index
        if row_index >= len(self.drug_entries):
            return

        # Get the drug entry dict at that row
        drug_entry = self.drug_entries[row_index]

        # Destroy each widget
        for widget_name in ['drug_name', 'dosage', 'breakfast', 'lunch', 'dinner',
                            'after_meals', 'before_meals', 'duration', 'instructions', 'remove_btn']:
            if widget_name in drug_entry:
                drug_entry[widget_name].destroy()

        # Remove the entry from the list
        del self.drug_entries[row_index]

        self.reposition_drug_entries()

    def save_prescription_to_excel(self):
        # Get patient information
        patient_name = self.patient_name_entry.get()
        date = self.cur_date_entry.get()
        age = self.patient_age_entry.get()
        sex = self.patient_sex_entry.get()
        diagnosis = self.patient_diagnosis_entry.get()

        # Collect all drug names and doses
        drug_names = []
        drug_doses = []

        for drug_entry in self.drug_entries:
            drug_name = drug_entry['drug_name'].get()
            dose = drug_entry['dosage'].get()

            if drug_name.strip():  # Only add if drug name is not empty
                drug_names.append(drug_name)
                drug_doses.append(dose)

        # Convert lists to arrays (as strings for Excel)
        Drugs = str(drug_names)  # This will create: ['Drug1', 'Drug2', 'Drug3']
        Doses = str(drug_doses)  # This will create: ['10mg', '20mg', '5ml']

        # Load Excel file
        path = "Sample.xlsx"
        wb = xl.load_workbook(path)
        sheet = wb.active

        # Create row with all data
        row_values = [patient_name, date, age, sex, diagnosis, Drugs, Doses]

        # Append to Excel
        sheet.append(row_values)
        wb.save(path)

        print(f"Prescription saved for {patient_name}")
        print(f"Drugs: {Drugs}")
        print(f"Doses: {Doses}")

    def reposition_drug_entries(self):
        for i, entry in enumerate(self.drug_entries):
            entry['drug_name'].grid(row=i, column=0, padx=5, pady=2)
            entry['dosage'].grid(row=i, column=1, padx=5, pady=2)
            entry['breakfast'].grid(row=i, column=2, padx=5, pady=2)
            entry['lunch'].grid(row=i, column=3, padx=5, pady=2)
            entry['dinner'].grid(row=i, column=4, padx=5, pady=2)
            entry['after_meals'].grid(row=i, column=5, padx=5, pady=2)
            entry['before_meals'].grid(row=i, column=6, padx=5, pady=2)
            entry['duration'].grid(row=i, column=7, padx=5, pady=2)
            entry['instructions'].grid(row=i, column=8, padx=5, pady=2)
            entry['remove_btn'].grid(row=i, column=9, padx=5, pady=2)

            # Update remove button command with new index
            entry['remove_btn'].config(command=lambda idx=i: self.remove_drug_entry(idx))
            entry['row'] = i

            # Re-bind events after repositioning
            entry['drug_name'].bind('<KeyRelease>', self.check)
            entry['drug_name'].bind('<FocusIn>', lambda e, widget=entry['drug_name']: self.set_current_widget(widget))

    def update(self, data):
        self.list_box.delete(0, tk.END)
        for value in data:
            self.list_box.insert(tk.END, value)

    def fillout(self, event):
        if self.current_search_widget:
            self.current_search_widget.delete(0, tk.END)
            self.current_search_widget.insert(0, self.list_box.get(ACTIVE))

    def check(self, event):
        # Get the widget that triggered the event
        widget = event.widget
        self.current_search_widget = widget

        typed = widget.get()
        if typed == "":
            data = self.drug_list[:20]  # show default top 20
        else:
            data = self.trie.search_prefix(typed, limit=20)
        self.update(data)

    def drug_list_view_frame(self):
        self.drug_list_frame = ttk.Frame(self.frame)
        self.drug_list_frame.grid(row=0, column=1, sticky='nsew')

        self.list_box = tk.Listbox(self.drug_list_frame, width=60)
        self.list_box.pack(expand=True, fill='both')

        # Load drug data
        try:
            df = pd.read_csv("Drugs.csv")
            self.drug_list = df['name'].tolist()
        except FileNotFoundError:
            # Create a sample drug list if file doesn't exist
            self.drug_list = ['Aspirin', 'Ibuprofen', 'Acetaminophen', 'Amoxicillin', 'Ciprofloxacin']
            print("Warning: Drugs.csv not found. Using sample drug list.")

        self.trie = Trie()
        for drug in self.drug_list:
            self.trie.insert(drug)
        self.update(self.drug_list)

        self.list_box.bind('<<ListboxSelect>>', self.fillout)


if __name__ == "__main__":
    window = tk.Tk()
    window.geometry("1920x1080")
    app = PrescriptionApp(window)
    window.mainloop()